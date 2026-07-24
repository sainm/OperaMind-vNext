import json
from pathlib import Path
from typing import Any

import pytest
from openpyxl import Workbook, load_workbook

from operamind.application import DocumentDiffBlockedError, DocumentDiffRequest, DocumentDiffService
from operamind.contracts import ContractCatalog
from operamind.domain.document_conventions import DocumentConvention
from operamind.infrastructure.documents import DocumentSignalExtractorRegistry
from operamind.profiles import ProfileCatalog

ROOT = Path(__file__).parents[2]


def load_screen_convention() -> DocumentConvention:
    profile: dict[str, Any] = json.loads(
        (ROOT / "profiles/screen-design-convention-profile.example.json").read_text(
            encoding="utf-8"
        )
    )
    ProfileCatalog.load(ROOT / "profiles").validate_profile(profile)
    return DocumentConvention.from_validated_profile(profile)


def load_api_convention() -> DocumentConvention:
    profile: dict[str, Any] = json.loads(
        (ROOT / "profiles/document-convention-profile.example.json").read_text(encoding="utf-8")
    )
    ProfileCatalog.load(ROOT / "profiles").validate_profile(profile)
    return DocumentConvention.from_validated_profile(profile)


def write_screen_design(path: Path, default_value: str) -> None:
    workbook = Workbook()
    overview = workbook.active
    overview.title = "画面概要"
    overview.append(["画面ID", "SCREEN_EXPENSE_LIST"])
    items = workbook.create_sheet("画面項目一覧")
    items.append(["項目名", "種別", "初期値", "備考"])
    items.append(
        [
            "expense-search-status",
            "セレクト",
            default_value,
            "ステータスフィルタ",
        ]
    )
    workbook.save(path)
    workbook.close()


def test_document_diff_service_returns_valid_change_envelope(tmp_path: Path) -> None:
    before_path = tmp_path / "02_画面設計書_before.xlsx"
    after_path = tmp_path / "02_画面設計書_after.xlsx"
    write_screen_design(before_path, "申請中")
    write_screen_design(after_path, "すべて")
    contracts = ContractCatalog.load(ROOT / "contracts")
    service = DocumentDiffService(
        extractors=DocumentSignalExtractorRegistry.default(),
        contracts=contracts,
    )

    result = service.run(
        DocumentDiffRequest(
            project_id="visiondemo",
            domain="ui",
            fact_type="screen_element",
            source_snapshot_id="snapshot-before",
            target_snapshot_id="snapshot-after",
            before_path=before_path,
            after_path=after_path,
        ),
        load_screen_convention(),
    )
    payload = result.to_payload()

    assert result.source_fact_count == result.target_fact_count == 1
    assert result.source_snapshot.snapshot_id == "snapshot-before"
    assert result.target_snapshot.snapshot_id == "snapshot-after"
    assert result.source_variant_id == result.target_variant_id == "screen-item-table-ja"
    assert result.source_ignored_sections == ("画面概要:below_auto_match_threshold",)
    assert result.target_ignored_sections == result.source_ignored_sections
    assert result.source_extractor_ref.startswith("operamind-xlsx-structural@1+openpyxl-")
    assert result.target_extractor_ref == result.source_extractor_ref
    assert len(result.source_content_digest) == len(result.target_content_digest) == 64
    assert len(result.changes) == 1
    assert result.changes[0].stable_key == (
        "screen_element:screen_expense_list/expense-search-status"
    )
    assert payload["structured_change_count"] == 1
    changes = payload["changes"]
    assert isinstance(changes, list)
    contracts.validate_artifact(changes[0])


def test_document_diff_maps_multiple_sheet_variants_into_one_snapshot(
    tmp_path: Path,
) -> None:
    before_path = tmp_path / "画面設計書_before.xlsx"
    after_path = tmp_path / "画面設計書_after.xlsx"

    def write(path: Path, default_value: str) -> None:
        workbook = Workbook()
        overview = workbook.active
        overview.title = "画面概要"
        overview.append(["画面名", "経費一覧"])
        overview.append(["画面ID", "SCREEN_EXPENSE_LIST"])
        items = workbook.create_sheet("画面項目一覧")
        items.append(["項目名", "種別", "初期値", "備考"])
        items.append(["expense-search-status", "セレクト", default_value, "ステータス"])
        events = workbook.create_sheet("イベント一覧")
        events.append(["No", "イベント名", "発生源", "トリガー", "処理内容"])
        events.append([1, "検索", "expense-search-btn", "クリック", "一覧を更新"])
        workbook.save(path)
        workbook.close()

    write(before_path, "申請中")
    write(after_path, "すべて")
    service = DocumentDiffService(
        extractors=DocumentSignalExtractorRegistry.default(),
        contracts=ContractCatalog.load(ROOT / "contracts"),
    )

    result = service.run(
        DocumentDiffRequest(
            project_id="visiondemo",
            domain="ui",
            fact_type="screen_element",
            source_snapshot_id="snapshot-before-multi",
            target_snapshot_id="snapshot-after-multi",
            before_path=before_path,
            after_path=after_path,
        ),
        load_screen_convention(),
    )

    assert result.source_fact_count == result.target_fact_count == 3
    assert result.source_snapshot_variant_ids == (
        "screen-overview-table-ja",
        "screen-item-table-ja",
        "screen-event-table-ja",
    )
    assert result.target_snapshot_variant_ids == result.source_snapshot_variant_ids
    assert {fact.fact.stable_key for fact in result.source_snapshot.facts} == {
        "screen_element:screen_expense_list/%E7%B5%8C%E8%B2%BB%E4%B8%80%E8%A6%A7",
        "screen_element:screen_expense_list/expense-search-status",
        "screen_element:screen_expense_list/%E6%A4%9C%E7%B4%A2",
    }
    assert len(result.changes) == 1
    assert result.changes[0].stable_key == (
        "screen_element:screen_expense_list/expense-search-status"
    )
    assert result.to_payload()["source_variant_ids"] == list(result.source_snapshot_variant_ids)


def test_document_diff_maps_api_list_and_detail_tables_with_fact_provenance(
    tmp_path: Path,
) -> None:
    before_path = tmp_path / "API詳細設計書_before.xlsx"
    after_path = tmp_path / "API詳細設計書_after.xlsx"

    def write(path: Path) -> None:
        workbook = Workbook()
        detail = workbook.active
        detail.title = "searchExpense"
        detail.append(["API: searchExpense"])
        detail.append(["GET /expense/api"])
        detail.append([])
        detail.append(["【リクエストパラメータ】"])
        detail.append(["No", "項目名", "データ型", "必須", "説明"])
        detail.append([1, "status", "String", "任意", "ステータス"])
        detail.append(["【レスポンス】"])
        detail.append(["No", "項目名", "データ型", "説明"])
        detail.append([1, "expenses", "List", "経費一覧"])
        api_list = workbook.create_sheet("API一覧")
        api_list.append(["API名", "URL", "HTTPメソッド"])
        api_list.append(["経費検索", "/expense/api", "GET"])
        workbook.save(path)
        workbook.close()

    write(before_path)
    write(after_path)
    result = DocumentDiffService(
        extractors=DocumentSignalExtractorRegistry.default(),
        contracts=ContractCatalog.load(ROOT / "contracts"),
    ).run(
        DocumentDiffRequest(
            project_id="visiondemo",
            domain="api",
            fact_type="api",
            source_snapshot_id="api-before",
            target_snapshot_id="api-after",
            before_path=before_path,
            after_path=after_path,
        ),
        load_api_convention(),
    )

    assert result.source_fact_count == result.target_fact_count == 3
    assert result.source_snapshot_variant_ids == (
        "api-object-table",
        "api-list-url",
    )
    assert set(dict(result.source_fact_variant_ids).values()) == {
        "api-object-table",
        "api-list-url",
    }
    assert result.changes == ()


def test_document_diff_blocks_structurally_related_unmatched_sheet(
    tmp_path: Path,
) -> None:
    before_path = tmp_path / "before.xlsx"
    after_path = tmp_path / "after.xlsx"

    def write(path: Path, default_value: str) -> None:
        write_screen_design(path, default_value)
        workbook = load_workbook(path)
        broken = workbook.create_sheet("名称変更されたイベント")
        broken.append(["No", "イベント名", "発生源", "トリガー", "処理内容"])
        broken.append([1, "検索", "expense-search-btn", "クリック", "一覧を更新"])
        workbook.save(path)
        workbook.close()

    write(before_path, "申請中")
    write(after_path, "すべて")

    with pytest.raises(DocumentDiffBlockedError, match="名称変更されたイベント"):
        DocumentDiffService(
            extractors=DocumentSignalExtractorRegistry.default(),
            contracts=ContractCatalog.load(ROOT / "contracts"),
        ).run(
            DocumentDiffRequest(
                project_id="visiondemo",
                domain="ui",
                fact_type="screen_element",
                source_snapshot_id="broken-before",
                target_snapshot_id="broken-after",
                before_path=before_path,
                after_path=after_path,
            ),
            load_screen_convention(),
        )


def test_document_diff_rejects_source_changed_during_extraction(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    before_path = tmp_path / "before.xlsx"
    after_path = tmp_path / "after.xlsx"
    write_screen_design(before_path, "申請中")
    write_screen_design(after_path, "すべて")
    service = DocumentDiffService(
        extractors=DocumentSignalExtractorRegistry.default(),
        contracts=ContractCatalog.load(ROOT / "contracts"),
    )
    digests = iter(("1" * 64, "2" * 64, "3" * 64, "2" * 64))
    monkeypatch.setattr(
        "operamind.application.document_diff._file_digest", lambda _path: next(digests)
    )

    with pytest.raises(DocumentDiffBlockedError, match="changed while"):
        service.run(
            DocumentDiffRequest(
                project_id="visiondemo",
                domain="ui",
                fact_type="screen_element",
                source_snapshot_id="snapshot-before",
                target_snapshot_id="snapshot-after",
                before_path=before_path,
                after_path=after_path,
            ),
            load_screen_convention(),
        )


def test_document_diff_rejects_same_source_and_target_path(tmp_path: Path) -> None:
    path = tmp_path / "same.xlsx"
    write_screen_design(path, "申請中")
    service = DocumentDiffService(
        extractors=DocumentSignalExtractorRegistry.default(),
        contracts=ContractCatalog.load(ROOT / "contracts"),
    )

    with pytest.raises(ValueError, match="paths must differ"):
        service.run(
            DocumentDiffRequest(
                project_id="visiondemo",
                domain="ui",
                fact_type="screen_element",
                source_snapshot_id="snapshot-before",
                target_snapshot_id="snapshot-after",
                before_path=path,
                after_path=path,
            ),
            load_screen_convention(),
        )

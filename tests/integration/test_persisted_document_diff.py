import json
import os
from pathlib import Path
from typing import Any
from uuid import uuid4

import psycopg
import pytest
from openpyxl import Workbook

from operamind.application import (
    DocumentDiffRequest,
    DocumentDiffService,
    PersistedDocumentDiffRequest,
    PersistedDocumentDiffService,
)
from operamind.contracts import ContractCatalog
from operamind.domain import ChangeReviewStatus
from operamind.infrastructure.documents import DocumentSignalExtractorRegistry
from operamind.infrastructure.postgres import (
    ArtifactRepository,
    CanonicalRepository,
    DocumentIngestionResultRepository,
    DocumentNodeRepository,
    MigrationCatalog,
    MigrationRunner,
    PersistenceConflictError,
    ProfileRepository,
    StructuredChangeReviewDecision,
    StructuredChangeReviewRepository,
)
from operamind.profiles import ProfileCatalog

ROOT = Path(__file__).parents[2]
DATABASE_URL = os.getenv("OPERAMIND_TEST_DATABASE_URL")

pytestmark = pytest.mark.integration


def load_profile() -> dict[str, Any]:
    value: object = json.loads(
        (ROOT / "profiles/screen-design-convention-profile.example.json").read_text(
            encoding="utf-8"
        )
    )
    assert isinstance(value, dict)
    return value


def write_screen_design(path: Path, default_value: str) -> None:
    workbook = Workbook()
    overview = workbook.active
    assert overview is not None
    overview.title = "画面概要"
    overview.append(["画面ID", "SCREEN_EXPENSE_LIST"])
    items = workbook.create_sheet("画面項目一覧")
    items.append(["項目名", "種別", "初期値", "備考"])
    items.append(
        [
            "expense-search-status",
            "セレクト",
            default_value,
            "ステータスフィルタ",
        ]
    )
    workbook.save(path)
    workbook.close()


def write_multi_sheet_screen_design(path: Path, default_value: str) -> None:
    workbook = Workbook()
    overview = workbook.active
    assert overview is not None
    overview.title = "画面概要"
    overview.append(["画面名", "経費精算申請一覧"])
    overview.append(["画面ID", "SCREEN_EXPENSE_LIST"])
    items = workbook.create_sheet("画面項目一覧")
    items.append(["項目名", "種別", "初期値", "備考"])
    items.append(
        [
            "expense-search-status",
            "セレクト",
            default_value,
            "ステータスフィルタ",
        ]
    )
    events = workbook.create_sheet("イベント一覧")
    events.append(["No", "イベント名", "発生源", "トリガー", "処理内容"])
    events.append([1, "検索", "expense-search-btn", "クリック", "一覧を更新"])
    workbook.save(path)
    workbook.close()


def insert_project_case(
    connection: psycopg.Connection[Any],
    *,
    project_id: str,
    case_id: str,
    suffix: str,
) -> None:
    with connection.cursor() as cursor:
        cursor.execute(
            "INSERT INTO projects (project_id, name) VALUES (%s, %s)",
            (project_id, "Persisted Diff integration test"),
        )
        cursor.execute(
            """
            INSERT INTO repositories (repository_id, project_id, remote_url)
            VALUES (%s, %s, %s)
            """,
            (f"repository-{suffix}", project_id, f"https://example.invalid/{suffix}.git"),
        )
        cursor.execute(
            """
            INSERT INTO repository_revisions (
                repository_revision_id, repository_id, commit_sha
            ) VALUES (%s, %s, %s)
            """,
            (f"revision-{suffix}", f"repository-{suffix}", suffix),
        )
        cursor.execute(
            """
            INSERT INTO analysis_cases (
                analysis_case_id, project_id, repository_revision_id, status
            ) VALUES (%s, %s, %s, 'ingesting')
            """,
            (case_id, project_id, f"revision-{suffix}"),
        )


def build_service(connection: psycopg.Connection[Any]) -> PersistedDocumentDiffService:
    contracts = ContractCatalog.load(ROOT / "contracts")
    return PersistedDocumentDiffService(
        connection=connection,
        document_diff=DocumentDiffService(
            extractors=DocumentSignalExtractorRegistry.default(),
            contracts=contracts,
        ),
        contracts=contracts,
        profiles=ProfileCatalog.load(ROOT / "profiles"),
    )


def build_request(
    *,
    suffix: str,
    project_id: str,
    case_id: str,
    before_path: Path,
    after_path: Path,
) -> PersistedDocumentDiffRequest:
    return PersistedDocumentDiffRequest(
        diff=DocumentDiffRequest(
            project_id=project_id,
            domain="ui",
            fact_type="screen_element",
            source_snapshot_id=f"snapshot-before-{suffix}",
            target_snapshot_id=f"snapshot-after-{suffix}",
            before_path=before_path,
            after_path=after_path,
        ),
        ingestion_batch_id=f"ingestion-{suffix}",
        analysis_case_id=case_id,
        document_id=f"document-{suffix}",
        logical_name="02_画面設計書_経費精算申請一覧.xlsx",
        source_document_version_id=f"document-version-before-{suffix}",
        target_document_version_id=f"document-version-after-{suffix}",
        source_ref=f"immutable://design-docs/{suffix}/before.xlsx",
        target_ref=f"immutable://design-docs/{suffix}/after.xlsx",
        profile_version_id=f"profile-{suffix}",
        profile_binding_key="document:screen_design",
        profile_activation_event_id=f"profile-activation-{suffix}",
        activated_by="reviewer@example.invalid",
        activation_reason="Reviewed P1 integration Profile",
    )


@pytest.mark.skipif(DATABASE_URL is None, reason="OPERAMIND_TEST_DATABASE_URL is not set")
def test_persisted_document_diff_is_atomic_and_idempotent(tmp_path: Path) -> None:
    assert DATABASE_URL is not None
    suffix = uuid4().hex
    project_id = f"project-{suffix}"
    case_id = f"case-{suffix}"
    before_path = tmp_path / "02_画面設計書_before.xlsx"
    after_path = tmp_path / "02_画面設計書_after.xlsx"
    write_screen_design(before_path, "申請中")
    write_screen_design(after_path, "すべて")
    request = build_request(
        suffix=suffix,
        project_id=project_id,
        case_id=case_id,
        before_path=before_path,
        after_path=after_path,
    )

    with psycopg.connect(DATABASE_URL) as connection:
        MigrationRunner(connection, MigrationCatalog.load(ROOT / "migrations")).apply()
        insert_project_case(
            connection,
            project_id=project_id,
            case_id=case_id,
            suffix=suffix,
        )
        service = build_service(connection)
        first = service.run(request, load_profile())
        replay = service.run(request, load_profile())

        assert replay.ingestion_artifact == first.ingestion_artifact
        assert replay.artifact_digests == first.artifact_digests
        ingestion = first.ingestion_artifact
        ContractCatalog.load(ROOT / "contracts").validate_artifact(ingestion)
        assert ingestion["document_profile_refs"] == ["screen-design-conventions-example@1.0.0"]
        assert ingestion["ingestion_result_event_id"] == first.initial_ingestion_event_id
        assert ingestion["document_profiles"] == [
            {
                "profile_version_id": request.profile_version_id,
                "binding_key": request.profile_binding_key,
                "activation_event_id": request.profile_activation_event_id,
                "profile_ref": "screen-design-conventions-example@1.0.0",
            }
        ]
        assert ingestion["source_content_digest"] == first.diff.source_content_digest
        assert ingestion["target_content_digest"] == first.diff.target_content_digest
        assert ingestion["source_extractor_ref"].startswith("operamind-xlsx-structural@1+openpyxl-")
        assert ingestion["target_extractor_ref"] == ingestion["source_extractor_ref"]
        assert ingestion["uploaded_document_count"] == 1
        assert ingestion["changed_document_count"] == 1
        assert ingestion["structured_change_count"] == 1
        assert ingestion["eligible_index_target_count"] == 1
        assert ingestion["indexed_target_count"] == 0
        assert ingestion["embedding_index_status"] == "not_started"
        assert ingestion["status"] == "needs_review"
        assert ingestion["blocking_reasons"] == [
            "structured_changes_require_review",
            "embedding_index_not_started",
        ]

        canonical = CanonicalRepository(connection, ContractCatalog.load(ROOT / "contracts"))
        assert (
            canonical.get_snapshot(
                project_id=project_id,
                snapshot_id=request.diff.source_snapshot_id,
            )
            == first.diff.source_snapshot
        )
        artifacts = ArtifactRepository(connection, ContractCatalog.load(ROOT / "contracts"))
        assert artifacts.get(request.ingestion_batch_id) == ingestion
        ingestion_events = DocumentIngestionResultRepository(
            connection,
            ContractCatalog.load(ROOT / "contracts"),
        )
        initial_event = ingestion_events.get_latest(
            project_id=project_id,
            ingestion_batch_id=request.ingestion_batch_id,
        )
        assert initial_event is not None
        assert initial_event.event_id == first.initial_ingestion_event_id
        assert initial_event.previous_event_id is None
        assert initial_event.artifact == ingestion
        nodes = DocumentNodeRepository(connection)
        assert nodes.list_indexable(
            project_id=project_id,
            snapshot_id=request.diff.target_snapshot_id,
        ) == tuple(node for node in first.target_nodes if node.index_eligible)
        profiles = ProfileRepository(connection, ProfileCatalog.load(ROOT / "profiles"))
        active = profiles.get_active(
            project_id=project_id,
            binding_key=request.profile_binding_key,
        )
        assert active is not None
        assert active.profile_version_id == request.profile_version_id

        with connection.cursor() as cursor:
            cursor.execute(
                """
                SELECT content_digest, extractor_ref
                FROM document_versions
                WHERE document_version_id IN (%s, %s)
                ORDER BY document_version_id
                """,
                (
                    request.source_document_version_id,
                    request.target_document_version_id,
                ),
            )
            document_versions = cursor.fetchall()
            assert {str(row[0]) for row in document_versions} == {
                first.diff.source_content_digest,
                first.diff.target_content_digest,
            }
            assert all(
                str(row[1]).startswith("operamind-xlsx-structural@1+openpyxl-")
                for row in document_versions
            )
        with connection.cursor() as cursor:
            cursor.execute(
                """
                SELECT
                    (SELECT count(*) FROM profile_activation_events WHERE project_id = %s),
                    (SELECT count(*) FROM document_snapshots WHERE project_id = %s),
                    (SELECT count(*) FROM document_facts WHERE project_id = %s),
                    (SELECT count(*) FROM document_nodes WHERE project_id = %s),
                    (SELECT count(*) FROM structured_changes WHERE project_id = %s),
                    (SELECT count(*) FROM artifact_records WHERE project_id = %s),
                    (SELECT count(*) FROM document_ingestion_result_events
                     WHERE project_id = %s)
                """,
                (
                    project_id,
                    project_id,
                    project_id,
                    project_id,
                    project_id,
                    project_id,
                    project_id,
                ),
            )
            assert cursor.fetchone() == (1, 2, 2, 4, 1, 2, 1)
        connection.rollback()


@pytest.mark.skipif(DATABASE_URL is None, reason="OPERAMIND_TEST_DATABASE_URL is not set")
def test_persisted_multi_sheet_diff_retains_snapshot_and_fact_variant_provenance(
    tmp_path: Path,
) -> None:
    assert DATABASE_URL is not None
    suffix = uuid4().hex
    project_id = f"project-{suffix}"
    case_id = f"case-{suffix}"
    before_path = tmp_path / "画面設計書_before.xlsx"
    after_path = tmp_path / "画面設計書_after.xlsx"
    write_multi_sheet_screen_design(before_path, "申請中")
    write_multi_sheet_screen_design(after_path, "すべて")
    request = build_request(
        suffix=suffix,
        project_id=project_id,
        case_id=case_id,
        before_path=before_path,
        after_path=after_path,
    )

    with psycopg.connect(DATABASE_URL) as connection:
        MigrationRunner(connection, MigrationCatalog.load(ROOT / "migrations")).apply()
        insert_project_case(
            connection,
            project_id=project_id,
            case_id=case_id,
            suffix=suffix,
        )
        result = build_service(connection).run(request, load_profile())

        expected_variants = (
            "screen-overview-table-ja",
            "screen-item-table-ja",
            "screen-event-table-ja",
        )
        assert result.diff.source_snapshot_variant_ids == expected_variants
        assert result.diff.target_snapshot_variant_ids == expected_variants
        assert set(dict(result.diff.source_fact_variant_ids).values()) == set(expected_variants)
        assert result.ingestion_artifact["source_variant_ids"] == list(expected_variants)
        assert result.ingestion_artifact["source_fact_variant_ids"] == dict(
            result.diff.source_fact_variant_ids
        )

        with connection.cursor() as cursor:
            cursor.execute(
                """
                SELECT selected_variant_id, selected_variant_ids
                FROM snapshot_memberships
                WHERE project_id = %s AND document_snapshot_id = %s
                """,
                (project_id, request.diff.source_snapshot_id),
            )
            membership = cursor.fetchone()
            cursor.execute(
                """
                SELECT selected_variant_id, count(*)
                FROM document_fact_variants
                WHERE project_id = %s AND document_snapshot_id = %s
                GROUP BY selected_variant_id
                ORDER BY selected_variant_id
                """,
                (project_id, request.diff.source_snapshot_id),
            )
            fact_variants = cursor.fetchall()

        assert membership == (expected_variants[0], list(expected_variants))
        assert fact_variants == [
            ("screen-event-table-ja", 1),
            ("screen-item-table-ja", 1),
            ("screen-overview-table-ja", 1),
        ]
        connection.rollback()


@pytest.mark.skipif(DATABASE_URL is None, reason="OPERAMIND_TEST_DATABASE_URL is not set")
def test_persisted_document_diff_rolls_back_when_case_has_wrong_project(
    tmp_path: Path,
) -> None:
    assert DATABASE_URL is not None
    suffix = uuid4().hex
    project_id = f"project-{suffix}"
    other_suffix = uuid4().hex
    other_project_id = f"project-{other_suffix}"
    other_case_id = f"case-{other_suffix}"
    before_path = tmp_path / "02_画面設計書_before.xlsx"
    after_path = tmp_path / "02_画面設計書_after.xlsx"
    write_screen_design(before_path, "申請中")
    write_screen_design(after_path, "すべて")
    request = build_request(
        suffix=suffix,
        project_id=project_id,
        case_id=other_case_id,
        before_path=before_path,
        after_path=after_path,
    )

    with psycopg.connect(DATABASE_URL) as connection:
        MigrationRunner(connection, MigrationCatalog.load(ROOT / "migrations")).apply()
        insert_project_case(
            connection,
            project_id=project_id,
            case_id=f"case-{suffix}",
            suffix=suffix,
        )
        insert_project_case(
            connection,
            project_id=other_project_id,
            case_id=other_case_id,
            suffix=other_suffix,
        )
        connection.commit()

        with pytest.raises(
            ValueError,
            match="Analysis case does not belong to the Artifact project",
        ):
            build_service(connection).run(request, load_profile())

        with connection.cursor() as cursor:
            cursor.execute(
                """
                SELECT
                    (SELECT count(*) FROM profile_versions WHERE profile_version_id = %s),
                    (SELECT count(*) FROM document_snapshots WHERE project_id = %s),
                    (SELECT count(*) FROM document_nodes WHERE project_id = %s),
                    (SELECT count(*) FROM structured_changes WHERE project_id = %s),
                    (SELECT count(*) FROM artifact_records WHERE project_id = %s),
                    (SELECT count(*) FROM document_ingestion_result_events
                     WHERE project_id = %s)
                """,
                (
                    request.profile_version_id,
                    project_id,
                    project_id,
                    project_id,
                    project_id,
                    project_id,
                ),
            )
            assert cursor.fetchone() == (0, 0, 0, 0, 0, 0)
        connection.rollback()


@pytest.mark.skipif(DATABASE_URL is None, reason="OPERAMIND_TEST_DATABASE_URL is not set")
def test_structured_change_review_is_append_only_idempotent_and_stale_safe(
    tmp_path: Path,
) -> None:
    assert DATABASE_URL is not None
    suffix = uuid4().hex
    project_id = f"project-{suffix}"
    case_id = f"case-{suffix}"
    other_suffix = uuid4().hex
    other_project_id = f"project-{other_suffix}"
    before_path = tmp_path / "02_画面設計書_before.xlsx"
    after_path = tmp_path / "02_画面設計書_after.xlsx"
    write_screen_design(before_path, "申請中")
    write_screen_design(after_path, "すべて")
    request = build_request(
        suffix=suffix,
        project_id=project_id,
        case_id=case_id,
        before_path=before_path,
        after_path=after_path,
    )

    with psycopg.connect(DATABASE_URL) as connection:
        MigrationRunner(connection, MigrationCatalog.load(ROOT / "migrations")).apply()
        insert_project_case(
            connection,
            project_id=project_id,
            case_id=case_id,
            suffix=suffix,
        )
        insert_project_case(
            connection,
            project_id=other_project_id,
            case_id=f"case-{other_suffix}",
            suffix=other_suffix,
        )
        result = build_service(connection).run(request, load_profile())
        change_id = result.diff.changes[0].change_id
        repository = StructuredChangeReviewRepository(connection)

        initial = repository.get_state(project_id=project_id, change_id=change_id)
        assert initial is not None
        assert initial.status is ChangeReviewStatus.NEEDS_REVIEW
        assert initial.review_event_id is None

        first_event_id = f"review-accepted-{suffix}"
        assert repository.review(
            review_event_id=first_event_id,
            project_id=project_id,
            change_id=change_id,
            decision=StructuredChangeReviewDecision.ACCEPTED,
            reviewed_by="reviewer@example.invalid",
            reason="Canonical change matches the source design",
            expected_previous_review_event_id=None,
        )
        assert not repository.review(
            review_event_id=first_event_id,
            project_id=project_id,
            change_id=change_id,
            decision=StructuredChangeReviewDecision.ACCEPTED,
            reviewed_by="reviewer@example.invalid",
            reason="Canonical change matches the source design",
            expected_previous_review_event_id=None,
        )
        with (
            pytest.raises(psycopg.errors.UniqueViolation),
            connection.transaction(),
            connection.cursor() as cursor,
        ):
            cursor.execute(
                """
                    INSERT INTO structured_change_review_events (
                        review_event_id,
                        project_id,
                        structured_change_id,
                        previous_review_event_id,
                        previous_review_status,
                        decision,
                        reviewed_by,
                        reason
                    ) VALUES (%s, %s, %s, NULL, 'needs_review', 'rejected', %s, %s)
                    """,
                (
                    f"review-forked-first-{suffix}",
                    project_id,
                    change_id,
                    "intruder@example.invalid",
                    "Attempted second first event",
                ),
            )
        with pytest.raises(PersistenceConflictError, match="different content"):
            repository.review(
                review_event_id=first_event_id,
                project_id=project_id,
                change_id=change_id,
                decision=StructuredChangeReviewDecision.ACCEPTED,
                reviewed_by="reviewer@example.invalid",
                reason="Conflicting replay reason",
                expected_previous_review_event_id=None,
            )
        with pytest.raises(ValueError, match="Stale StructuredChange review"):
            repository.review(
                review_event_id=f"review-stale-{suffix}",
                project_id=project_id,
                change_id=change_id,
                decision=StructuredChangeReviewDecision.REJECTED,
                reviewed_by="second-reviewer@example.invalid",
                reason="Stale browser decision",
                expected_previous_review_event_id=None,
            )
        with pytest.raises(ValueError, match="must change the effective status"):
            repository.review(
                review_event_id=f"review-redundant-{suffix}",
                project_id=project_id,
                change_id=change_id,
                decision=StructuredChangeReviewDecision.ACCEPTED,
                reviewed_by="second-reviewer@example.invalid",
                reason="Redundant acceptance",
                expected_previous_review_event_id=first_event_id,
            )
        with pytest.raises(ValueError, match="does not belong to the review project"):
            repository.review(
                review_event_id=f"review-cross-project-{suffix}",
                project_id=other_project_id,
                change_id=change_id,
                decision=StructuredChangeReviewDecision.REJECTED,
                reviewed_by="intruder@example.invalid",
                reason="Cross-project decision",
                expected_previous_review_event_id=first_event_id,
            )

        second_event_id = f"review-rejected-{suffix}"
        assert repository.review(
            review_event_id=second_event_id,
            project_id=project_id,
            change_id=change_id,
            decision=StructuredChangeReviewDecision.REJECTED,
            reviewed_by="lead-reviewer@example.invalid",
            reason="Acceptance was reversed after source re-check",
            expected_previous_review_event_id=first_event_id,
        )
        with (
            pytest.raises(psycopg.errors.UniqueViolation),
            connection.transaction(),
            connection.cursor() as cursor,
        ):
            cursor.execute(
                """
                    INSERT INTO structured_change_review_events (
                        review_event_id,
                        project_id,
                        structured_change_id,
                        previous_review_event_id,
                        previous_review_status,
                        decision,
                        reviewed_by,
                        reason
                    ) VALUES (%s, %s, %s, %s, 'accepted', 'rejected', %s, %s)
                    """,
                (
                    f"review-forked-successor-{suffix}",
                    project_id,
                    change_id,
                    first_event_id,
                    "intruder@example.invalid",
                    "Attempted second successor",
                ),
            )
        assert not repository.review(
            review_event_id=first_event_id,
            project_id=project_id,
            change_id=change_id,
            decision=StructuredChangeReviewDecision.ACCEPTED,
            reviewed_by="reviewer@example.invalid",
            reason="Canonical change matches the source design",
            expected_previous_review_event_id=None,
        )

        final = repository.get_state(project_id=project_id, change_id=change_id)
        assert final is not None
        assert final.status is ChangeReviewStatus.REJECTED
        assert final.review_event_id == second_event_id
        assert repository.list_states(
            project_id=project_id,
            source_snapshot_id=request.diff.source_snapshot_id,
            target_snapshot_id=request.diff.target_snapshot_id,
        ) == (final,)
        with connection.cursor() as cursor:
            cursor.execute(
                """
                SELECT review_status
                FROM structured_changes
                WHERE structured_change_id = %s
                """,
                (change_id,),
            )
            assert cursor.fetchone() == ("needs_review",)
            cursor.execute(
                """
                SELECT previous_review_event_id, previous_review_status, decision
                FROM structured_change_review_events
                WHERE structured_change_id = %s
                ORDER BY review_sequence
                """,
                (change_id,),
            )
            assert cursor.fetchall() == [
                (None, "needs_review", "accepted"),
                (first_event_id, "accepted", "rejected"),
            ]
        connection.rollback()

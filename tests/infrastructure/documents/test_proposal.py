from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from operamind.infrastructure.documents import DocumentCellChange, XlsxDocumentProposalWriter


def _source(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "項目定義"
    sheet["F11"] = "下書き"
    workbook.save(path)
    workbook.close()


def _change() -> DocumentCellChange:
    return DocumentCellChange(
        operation_id="operation-1",
        document="screen.xlsx",
        sheet="項目定義",
        cell="F11",
        field="description",
        before="下書き",
        after="下書き、差戻し",
        source_ref="screen.xlsx#項目定義!F11",
    )


def test_xlsx_proposal_applies_exact_cell_without_mutating_source(tmp_path: Path) -> None:
    source = tmp_path / "screen.xlsx"
    target = tmp_path / "generated" / "screen-after.xlsx"
    _source(source)

    result = XlsxDocumentProposalWriter().apply(
        source_path=source,
        target_path=target,
        changes=(_change(),),
    )

    source_book = load_workbook(source, read_only=True)
    target_book = load_workbook(target, read_only=True)
    try:
        assert source_book["項目定義"]["F11"].value == "下書き"
        assert target_book["項目定義"]["F11"].value == "下書き、差戻し"
    finally:
        source_book.close()
        target_book.close()
    assert result.source_content_digest != result.target_content_digest


def test_xlsx_proposal_rejects_precondition_mismatch_and_removes_temp(tmp_path: Path) -> None:
    source = tmp_path / "screen.xlsx"
    target = tmp_path / "screen-after.xlsx"
    _source(source)
    change = _change()
    wrong = DocumentCellChange(
        operation_id=change.operation_id,
        document=change.document,
        sheet=change.sheet,
        cell=change.cell,
        field=change.field,
        before="unexpected",
        after=change.after,
        source_ref=change.source_ref,
    )

    with pytest.raises(ValueError, match="precondition mismatch"):
        XlsxDocumentProposalWriter().apply(
            source_path=source,
            target_path=target,
            changes=(wrong,),
        )

    assert not target.exists()
    assert not (tmp_path / ".screen-after.operamind-tmp.xlsx").exists()

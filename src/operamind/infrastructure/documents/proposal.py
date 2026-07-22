"""Deterministic, cell-bounded XLSX proposal generation."""

from __future__ import annotations

import re
import shutil
from dataclasses import dataclass
from hashlib import sha256
from pathlib import Path

from openpyxl import load_workbook

_SOURCE_REF = re.compile(r"^(?P<document>[^#]+)#(?P<sheet>[^!]+)!(?P<cell>[A-Z]+[1-9][0-9]*)$")


@dataclass(frozen=True, slots=True)
class DocumentCellChange:
    """One exact expected cell transition derived from an approved change."""

    operation_id: str
    document: str
    sheet: str
    cell: str
    field: str
    before: object
    after: object
    source_ref: str

    @classmethod
    def from_field_delta(cls, *, operation_id: str, delta: dict[str, object]) -> DocumentCellChange:
        source_ref = str(delta.get("source_ref", ""))
        match = _SOURCE_REF.fullmatch(source_ref)
        if match is None:
            raise ValueError(f"Document field delta has an invalid source_ref: {source_ref}")
        field = str(delta.get("field", ""))
        if not operation_id.strip() or not field.strip():
            raise ValueError("Document proposal operation identity and field must not be blank")
        return cls(
            operation_id=operation_id,
            document=match.group("document"),
            sheet=match.group("sheet"),
            cell=match.group("cell"),
            field=field,
            before=delta.get("before"),
            after=delta.get("after"),
            source_ref=source_ref,
        )

    def to_artifact(self) -> dict[str, object]:
        return {
            "operation_id": self.operation_id,
            "sheet": self.sheet,
            "cell": self.cell,
            "field": self.field,
            "before": self.before,
            "after": self.after,
            "source_ref": self.source_ref,
        }


@dataclass(frozen=True, slots=True)
class DocumentProposalWriteResult:
    source_content_digest: str
    target_content_digest: str
    target_path: Path


class XlsxDocumentProposalWriter:
    """Copy one workbook and apply only precondition-checked cell transitions."""

    def apply(
        self,
        *,
        source_path: Path,
        target_path: Path,
        changes: tuple[DocumentCellChange, ...],
    ) -> DocumentProposalWriteResult:
        source = source_path.resolve(strict=True)
        target = target_path.absolute()
        if source.suffix.casefold() != ".xlsx" or target.suffix.casefold() != ".xlsx":
            raise ValueError("Document proposal currently supports XLSX only")
        if source == target.resolve():
            raise ValueError("Document proposal target must differ from the source")
        if not changes:
            raise ValueError("Document proposal requires at least one cell change")
        if len({(change.sheet, change.cell) for change in changes}) != len(changes):
            raise ValueError("Document proposal cannot write the same cell twice")
        if any(change.document != source.name for change in changes):
            raise ValueError("Document proposal operation references another document")

        target.parent.mkdir(parents=True, exist_ok=True)
        temporary = target.with_name(f".{target.stem}.operamind-tmp.xlsx")
        shutil.copyfile(source, temporary)
        try:
            workbook = load_workbook(temporary, read_only=False, data_only=False, keep_links=False)
            try:
                for change in changes:
                    if change.sheet not in workbook.sheetnames:
                        raise ValueError(f"Document proposal sheet does not exist: {change.sheet}")
                    cell = workbook[change.sheet][change.cell]
                    if cell.value != change.before:
                        raise ValueError(
                            "Document proposal precondition mismatch: "
                            f"{change.sheet}!{change.cell} expected={change.before!r} "
                            f"actual={cell.value!r}"
                        )
                    cell.value = change.after
                workbook.save(temporary)
            finally:
                workbook.close()
        except Exception:
            temporary.unlink(missing_ok=True)
            raise
        temporary.replace(target)
        return DocumentProposalWriteResult(
            source_content_digest=_file_digest(source),
            target_content_digest=_file_digest(target),
            target_path=target,
        )

    def verify_pair(
        self,
        *,
        source_path: Path,
        target_path: Path,
        changes: tuple[DocumentCellChange, ...],
    ) -> DocumentProposalWriteResult:
        """Verify that an existing before/after pair contains the exact transitions."""

        source = source_path.resolve(strict=True)
        target = target_path.resolve(strict=True)
        if source == target or not changes:
            raise ValueError("Document pair must use distinct files and at least one change")
        source_book = load_workbook(source, read_only=True, data_only=False, keep_links=False)
        target_book = load_workbook(target, read_only=True, data_only=False, keep_links=False)
        try:
            for change in changes:
                if (
                    change.sheet not in source_book.sheetnames
                    or change.sheet not in target_book.sheetnames
                ):
                    raise ValueError(f"Document pair sheet does not exist: {change.sheet}")
                source_value = source_book[change.sheet][change.cell].value
                target_value = target_book[change.sheet][change.cell].value
                if source_value != change.before or target_value != change.after:
                    raise ValueError(
                        "Document pair does not contain the approved transition: "
                        f"{change.sheet}!{change.cell} "
                        f"expected={change.before!r}->{change.after!r} "
                        f"actual={source_value!r}->{target_value!r}"
                    )
        finally:
            source_book.close()
            target_book.close()
        return DocumentProposalWriteResult(
            source_content_digest=_file_digest(source),
            target_content_digest=_file_digest(target),
            target_path=target,
        )


def _file_digest(path: Path) -> str:
    digest = sha256()
    with path.open("rb") as stream:
        while chunk := stream.read(1024 * 1024):
            digest.update(chunk)
    return digest.hexdigest()
